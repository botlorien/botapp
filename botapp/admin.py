import csv
from django.http import HttpResponse
from django.contrib import admin
from .models import Bot, Task, TaskLog
from django.utils.translation import gettext_lazy as _
from django.utils.timezone import localtime
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.template.response import TemplateResponse
from django.urls import path
from django.db.models import Count
from rangefilter.filters import DateRangeFilter
from django.utils.html import format_html
from django.db.models import Avg, functions
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template

def export_tasklog_pdf(modeladmin, request, queryset):
    from datetime import datetime

    queryset = queryset.order_by('-start_time')[:100]
    context = {
        'logs': queryset,
        'generated_at': datetime.now(),
    }

    template = get_template("admin/botapp/tasklog/report_pdf.html")
    html = template.render(context)

    result = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=result)

    if pisa_status.err:
        return HttpResponse("Erro ao gerar PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="tasklog_report.pdf"'
    return response


export_tasklog_pdf.short_description = "📄 Exportar relatório em PDF"


def export_as_excel(modeladmin, request, queryset):
    model = modeladmin.model
    meta = model._meta

    # Campos padrão + campos relacionados personalizados
    field_names = [
        'task__bot__name',
        'task__bot__department',
        'task__name',
        'status',
        'start_time',
        'end_time',
        'duration',
        'user_login',
        'host_ip',
        'host_name',
        'os_platform',
        'exception_type',
        'env',
        'pid',
        'manual_trigger',
        'trigger_source',
    ]

    header_map = {
        'task__bot__name': "Bot",
        'task__bot__department': "Departamento",
        'task__name': "Tarefa",
        'start_time': "Início",
        'end_time': "Fim",
        'user_login': "Usuário",
        'host_ip': "IP da Máquina",
        'host_name': "Hostname",
        'os_platform': "Sistema",
        'exception_type': "Erro",
        'manual_trigger': "Execução Manual?",
        'trigger_source': "Origem",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{meta.verbose_name_plural}'

    # Cabeçalhos
    for col_num, field in enumerate(field_names, 1):
        col_letter = get_column_letter(col_num)
        label = header_map.get(field, field.replace('__', ' ').title())
        ws[f'{col_letter}1'] = label

    # Dados
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(field_names, 1):
            value = obj
            for part in field.split('__'):
                value = getattr(value, part, '')
                if value is None:
                    break
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={meta.verbose_name_plural}.xlsx'
    wb.save(response)
    return response


export_as_excel.short_description = "📥 Exportar para Excel (.xlsx)"


@admin.register(Bot)
class BotAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'version', 'department', 'is_active',
        'silence_threshold_minutes', 'silence_threshold_hours',
        'last_execution_at', 'created_at',
    )
    list_filter = ('is_active', 'department')
    search_fields = ('name', 'description', 'version')
    ordering = ('-updated_at',)
    readonly_fields = ('created_at', 'updated_at', 'last_execution_at', 'last_status')
    list_editable = ('silence_threshold_minutes', 'silence_threshold_hours')
    actions = [export_as_excel]


@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = ('name', 'bot', 'expected_duration_seconds', 'created_at', 'updated_at')
    list_filter = ('bot__department',)
    search_fields = ('name', 'description', 'bot__name')
    ordering = ('-updated_at',)
    readonly_fields = ('created_at', 'updated_at')
    list_editable = ('expected_duration_seconds',)
    actions = [export_as_excel]


@admin.register(TaskLog)
class TaskLogAdmin(admin.ModelAdmin):
    list_display = (
        'get_bot_name', 'get_bot_department', 'get_task_name',
        'status', 'start_time', 'end_time', 'duration',
        'user_login', 'host_name', 'os_platform', 'exception_type'
    )
    list_filter = (
        ('start_time', DateRangeFilter),
        'status',
        'task__bot__name',
        'task__bot__department',
        'env',
    )
    search_fields = (
        'task__name', 'task__bot__name', 'task__bot__department',
        'user_login', 'host_name', 'exception_type', 'error_message'
    )
    readonly_fields = [f.name for f in TaskLog._meta.fields]
    ordering = ('-start_time',)
    date_hierarchy = 'start_time'
    actions = [
        export_as_excel,
        export_tasklog_pdf,
    ]

    def has_add_permission(self, request):
        return False

    def get_task_name(self, obj):
        return obj.task.name
    get_task_name.short_description = "Tarefa"

    def get_bot_name(self, obj):
        return obj.task.bot.name
    get_bot_name.short_description = "Bot"

    def get_bot_department(self, obj):
        return obj.task.bot.department
    get_bot_department.short_description = "Departamento"


# ── Integração de CI ───────────────────────────────────────────────────────
from .models import CIConnection, CIJob, CIPipeline, CIProject, CISchedule  # noqa: E402


@admin.register(CIConnection)
class CIConnectionAdmin(admin.ModelAdmin):
    list_display = ('name', 'kind', 'base_url', 'namespace', 'enabled',
                    'last_sync_status', 'last_sync_at')
    list_filter = ('kind', 'enabled', 'last_sync_status')
    search_fields = ('name', 'base_url', 'namespace')
    # token_encrypted fica FORA de qualquer formulário: o valor não deve ser
    # legível nem editável pela UI. Ver docs/ci-integration-design.md §5.
    exclude = ('token_encrypted',)
    readonly_fields = ('last_discovery_at', 'last_sync_at', 'last_sync_status',
                       'last_sync_error', 'created_at', 'updated_at')


@admin.register(CIProject)
class CIProjectAdmin(admin.ModelAdmin):
    list_display = ('path', 'connection', 'monitored', 'scan_logs', 'bot',
                    'last_pipeline_status', 'last_pipeline_at', 'archived')
    list_filter = ('connection', 'monitored', 'scan_logs', 'archived',
                   'last_pipeline_status')
    search_fields = ('path', 'name')
    list_select_related = ('connection', 'bot')
    autocomplete_fields = ('bot',)
    readonly_fields = ('external_id', 'pipelines_cursor', 'last_pipeline_at',
                       'last_pipeline_status', 'last_pipeline_external_id',
                       'last_sync_error')


@admin.register(CISchedule)
class CIScheduleAdmin(admin.ModelAdmin):
    list_display = ('project', 'cron', 'cron_timezone', 'active', 'next_run_at')
    list_filter = ('active',)
    search_fields = ('project__path', 'description')
    list_select_related = ('project',)


@admin.register(CIPipeline)
class CIPipelineAdmin(admin.ModelAdmin):
    list_display = ('external_id', 'project', 'status', 'source',
                    'has_masked_error', 'created_at', 'duration')
    list_filter = ('status', 'source', 'has_masked_error')
    search_fields = ('project__path', 'ref', 'sha')
    list_select_related = ('project',)
    date_hierarchy = 'created_at'


@admin.register(CIJob)
class CIJobAdmin(admin.ModelAdmin):
    list_display = ('name', 'pipeline', 'status', 'stage', 'runner_description',
                    'duration')
    list_filter = ('status', 'stage')
    search_fields = ('name', 'pipeline__project__path')
    list_select_related = ('pipeline', 'pipeline__project')
